from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for
from flask_cors import CORS
import re
import unicodedata
import csv
import os
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
from functools import wraps
import msal
import uuid

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', os.urandom(24).hex())  # Change this in production!
CORS(app, supports_credentials=True)

# ---- Configuration ----
# Get the directory where this script is located
APP_ROOT = Path(__file__).parent.absolute()
CSV_DATA_DIR = APP_ROOT / "Data"

# OneDrive URLs (if set, will download from OneDrive instead of local files)
ONEDRIVE_EMPLOYEES_URL = os.environ.get('ONEDRIVE_EMPLOYEES_URL', '')
ONEDRIVE_PROJECTS_URL = os.environ.get('ONEDRIVE_PROJECTS_URL', '')
ONEDRIVE_PROJECT_EMPLOYEES_URL = os.environ.get('ONEDRIVE_PROJECT_EMPLOYEES_URL', '')

# Local CSV paths (fallback if OneDrive URLs not set)
EMPLOYEES_CSV = CSV_DATA_DIR / "employees.csv"
PROJECTS_CSV = CSV_DATA_DIR / "projects.csv"
PROJECT_EMPLOYEES_CSV = CSV_DATA_DIR / "project_employees.csv"

# ---- Azure AD Configuration ----
# Get these values from Azure Portal -> App Registrations -> Your App
AZURE_CLIENT_ID = os.environ.get('AZURE_CLIENT_ID', '')  # Application (client) ID
AZURE_CLIENT_SECRET = os.environ.get('AZURE_CLIENT_SECRET', '')  # Client secret value
AZURE_TENANT_ID = os.environ.get('AZURE_TENANT_ID', '')  # Directory (tenant) ID
AZURE_AUTHORITY = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}" if AZURE_TENANT_ID else None
AZURE_REDIRECT_PATH = "/getAToken"  # Used to form an absolute URL
AZURE_SCOPE = ["User.Read"]  # Permission scope
SESSION_TYPE = "filesystem"  # Use filesystem session storage

# ---- CSV Data Storage ----
employees_data = []
projects_data = []
project_employees_data = []

def download_from_onedrive(url, filename):
    """Download a file from OneDrive/SharePoint URL"""
    try:
        import requests
        print(f"Downloading {filename} from OneDrive...")
        
        # Convert SharePoint link to direct download if needed
        download_url = url
        if 'web=1' in url:
            download_url = url.replace('web=1', 'download=1')
        elif 'download=1' not in url and 'web=1' not in url:
            # If neither parameter exists, add download=1
            separator = '&' if '?' in url else '?'
            download_url = f"{url}{separator}download=1"
        
        print(f"Download URL: {download_url}")
        
        # Create session with headers to handle SharePoint authentication
        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        response = session.get(download_url, timeout=60, allow_redirects=True)
        response.raise_for_status()
        
        # Save to temporary location
        temp_dir = Path('/tmp') if Path('/tmp').exists() else CSV_DATA_DIR
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_file = temp_dir / filename
        
        with open(temp_file, 'wb') as f:
            f.write(response.content)
        
        print(f"Downloaded {filename} to {temp_file} ({len(response.content)} bytes)")
        return temp_file
    except Exception as e:
        print(f"Error downloading {filename} from OneDrive: {e}")
        import traceback
        traceback.print_exc()
        return None

def load_csv_data():
    """Load all CSV data into memory"""
    global employees_data, projects_data, project_employees_data
    
    # Determine CSV source (OneDrive or local)
    use_onedrive = bool(ONEDRIVE_EMPLOYEES_URL or ONEDRIVE_PROJECTS_URL or ONEDRIVE_PROJECT_EMPLOYEES_URL)
    
    if use_onedrive:
        print("Loading CSV data from OneDrive...")
        # Download from OneDrive if URLs are provided
        employees_file = download_from_onedrive(ONEDRIVE_EMPLOYEES_URL, 'employees.csv') if ONEDRIVE_EMPLOYEES_URL else EMPLOYEES_CSV
        projects_file = download_from_onedrive(ONEDRIVE_PROJECTS_URL, 'projects.csv') if ONEDRIVE_PROJECTS_URL else PROJECTS_CSV
        project_employees_file = download_from_onedrive(ONEDRIVE_PROJECT_EMPLOYEES_URL, 'project_employees.csv') if ONEDRIVE_PROJECT_EMPLOYEES_URL else PROJECT_EMPLOYEES_CSV
    else:
        print(f"Loading CSV data from local files: {CSV_DATA_DIR.absolute()}")
        employees_file = EMPLOYEES_CSV
        projects_file = PROJECTS_CSV
        project_employees_file = PROJECT_EMPLOYEES_CSV
    
    # Load employees
    try:
        if employees_file and (isinstance(employees_file, Path) and employees_file.exists() or isinstance(employees_file, str)):
            print(f"Reading employees from: {employees_file}")
            with open(employees_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                employees_data = list(reader)
                print(f"Read {len(employees_data)} rows from employees.csv")
                # Convert id to int for consistency, also ensure EmployeeID matches
                for emp in employees_data:
                    try:
                        # Try to get id field
                        emp_id = emp.get('id')
                        if not emp_id:
                            # Fallback to EmployeeID if id is not available
                            emp_id = emp.get('EmployeeID')
                        emp['id'] = int(emp_id) if emp_id else 0
                        # Also set EmployeeID to match for consistency
                        if not emp.get('EmployeeID'):
                            emp['EmployeeID'] = emp['id']
                    except (ValueError, TypeError):
                        emp['id'] = 0
                        emp['EmployeeID'] = 0
            print(f"Loaded {len(employees_data)} employees from CSV")
        else:
            print(f"ERROR: Employees CSV not found!")
            employees_data = []
    except Exception as e:
        print(f"ERROR loading employees CSV: {e}")
        import traceback
        traceback.print_exc()
        employees_data = []
    
    # Load projects
    try:
        if projects_file and (isinstance(projects_file, Path) and projects_file.exists() or isinstance(projects_file, str)):
            print(f"Reading projects from: {projects_file}")
            with open(projects_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                projects_data = list(reader)
                # Convert id to int for consistency
                for proj in projects_data:
                    try:
                        proj['id'] = int(proj.get('id', 0))
                    except (ValueError, TypeError):
                        proj['id'] = 0
            print(f"Loaded {len(projects_data)} projects from CSV")
        else:
            print(f"Warning: Projects CSV not found")
            projects_data = []
    except Exception as e:
        print(f"ERROR loading projects CSV: {e}")
        import traceback
        traceback.print_exc()
        projects_data = []
    
    # Load project-employee relationships
    try:
        if project_employees_file and (isinstance(project_employees_file, Path) and project_employees_file.exists() or isinstance(project_employees_file, str)):
            print(f"Reading project-employees from: {project_employees_file}")
            with open(project_employees_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                project_employees_data = list(reader)
                # Convert IDs to int
                for rel in project_employees_data:
                    try:
                        rel['ProjectID'] = int(rel.get('ProjectID', 0))
                        rel['EmployeeID'] = int(rel.get('EmployeeID', 0))
                    except (ValueError, TypeError):
                        rel['ProjectID'] = 0
                        rel['EmployeeID'] = 0
            print(f"Loaded {len(project_employees_data)} project-employee relationships from CSV")
        else:
            print(f"Warning: Project-employees CSV not found")
            project_employees_data = []
    except Exception as e:
        print(f"ERROR loading project-employees CSV: {e}")
        import traceback
        traceback.print_exc()
        project_employees_data = []

# Load CSV data on startup (only if not in serverless environment)
# In serverless, data will be loaded on first request via ensure_data_loaded()
try:
    # Only load if we have OneDrive URLs or local files exist
    # This prevents crashes in serverless environments where files don't exist yet
    if ONEDRIVE_EMPLOYEES_URL or ONEDRIVE_PROJECTS_URL or ONEDRIVE_PROJECT_EMPLOYEES_URL:
        # OneDrive URLs are set, try to load
        load_csv_data()
    elif EMPLOYEES_CSV.exists() and PROJECTS_CSV.exists() and PROJECT_EMPLOYEES_CSV.exists():
        # Local files exist, try to load
        load_csv_data()
    else:
        # No data source available yet, will load on first request
        print("No CSV data source available at startup. Data will be loaded on first request.")
except Exception as e:
    # Don't crash on import - data will be loaded on first request
    print(f"Could not load CSV data at startup: {e}")
    print("Data will be loaded on first request via ensure_data_loaded()")

def ensure_data_loaded():
    """Ensure CSV data is loaded (handles Flask debug reloader issue)"""
    global employees_data, projects_data, project_employees_data
    import os
    
    # Check if data needs to be loaded
    if not employees_data or len(employees_data) == 0:
        print("=" * 60)
        print("Data not loaded, reloading CSV files...")
        print(f"Current working directory: {os.getcwd()}")
        print(f"CSV directory path: {CSV_DATA_DIR}")
        print(f"CSV directory absolute path: {CSV_DATA_DIR.absolute()}")
        print(f"CSV directory exists: {CSV_DATA_DIR.exists()}")
        print(f"Employees CSV exists: {EMPLOYEES_CSV.exists()}")
        print(f"Employees CSV path: {EMPLOYEES_CSV.absolute()}")
        print("=" * 60)
        
        try:
            load_csv_data()
            print(f"After reload - employees_data length: {len(employees_data) if employees_data else 0}")
            print(f"After reload - projects_data length: {len(projects_data) if projects_data else 0}")
        except Exception as e:
            print(f"ERROR loading CSV data: {e}")
            import traceback
            traceback.print_exc()

# ---- Strong whitespace/encoding sanitizers ----
_PCT_WS = re.compile(r"(?i)(%20|%09|%0A|%0D|%C2%A0)+")

def _is_ws_like(ch: str) -> bool:
    if not ch:
        return False
    cat = unicodedata.category(ch)
    return ch.isspace() or cat in ("Zs", "Zl", "Zp", "Cf", "Cc")

def _clean_ws_like(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", s)
    return "".join(ch for ch in s if not _is_ws_like(ch))

def _sanitize_url(url: str) -> str:
    if not url:
        return url
    s = unicodedata.normalize("NFKC", url)
    s = _PCT_WS.sub("", s)

    parts = urlsplit(s)
    scheme = _clean_ws_like(parts.scheme) or "https"
    netloc = _clean_ws_like(parts.netloc)

    segs = parts.path.split("/")
    segs = [_clean_ws_like(seg) for seg in segs]
    path = "/".join(segs)

    # Handle spaces in filenames - encode them as %20 instead of removing
    path = re.sub(r"\s+(?=\.)", "%20", path)  # Space before file extension
    path = re.sub(r"(?<=\.)\s+", "%20", path)  # Space after file extension  
    path = re.sub(r"\s+", "%20", path)  # All other spaces

    query = _clean_ws_like(parts.query)
    fragment = _clean_ws_like(parts.fragment)

    sanitized = urlunsplit((scheme, netloc, path, query, fragment))
    return "".join(ch for ch in sanitized if not _is_ws_like(ch))

def _build_openasset_url(size_obj: dict):
    root_raw = (size_obj.get("http_root") or "")
    rel_raw  = (size_obj.get("http_relative_path") or "")

    root = _clean_ws_like(root_raw)
    # Don't remove whitespace from relative path - let _sanitize_url handle it properly
    rel = rel_raw.strip() if rel_raw else ""

    if not root or not rel:
        return None

    if root.startswith("//"):
        base = "https:" + root
    elif root.startswith(("http://", "https://")):
        base = root
    else:
        base = "https://" + root.lstrip("/")

    final_url = base.rstrip("/") + "/" + rel.lstrip("/")
    return _sanitize_url(final_url)

# ---- Azure AD Authentication Functions ----
def _build_msal_app(cache=None, authority=None):
    """Build MSAL application"""
    if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        return None
    try:
        return msal.ConfidentialClientApplication(
            AZURE_CLIENT_ID,
            authority=authority or AZURE_AUTHORITY,
            client_credential=AZURE_CLIENT_SECRET if AZURE_CLIENT_SECRET else None,
            token_cache=cache,
        )
    except Exception as e:
        print(f"Error building MSAL app: {e}")
        return None

def _get_token_from_cache(scope=None):
    """Get token from cache"""
    if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        return None
    try:
        cache = _load_cache()
        cca = _build_msal_app(cache=cache)
        if not cca:
            return None
        accounts = cca.get_accounts()
        if accounts:
            result = cca.acquire_token_silent(scope or AZURE_SCOPE, account=accounts[0])
            _save_cache(cache)
            return result
    except Exception as e:
        print(f"Error getting token from cache: {e}")
    return None

def _load_cache():
    """Load token cache from session"""
    if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        return None
    try:
        cache = msal.SerializableTokenCache()
        if session.get("token_cache"):
            cache.deserialize(session["token_cache"])
        return cache
    except Exception as e:
        print(f"Error loading cache: {e}")
        return None

def _save_cache(cache):
    """Save token cache to session"""
    if not cache or not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        return
    try:
        if cache.has_state_changed:
            session["token_cache"] = cache.serialize()
    except Exception as e:
        print(f"Error saving cache: {e}")

def login_required(f):
    """Decorator to require authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if Azure AD is configured
        if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
            # If not configured, allow access (for development)
            return f(*args, **kwargs)
        
        # Check if user is authenticated
        token = _get_token_from_cache()
        if not token:
            # Automatically redirect to login (no button needed)
            return redirect(url_for('login'))
        
        # Check if token is valid
        if "error" in token:
            return redirect(url_for('login'))
        
        return f(*args, **kwargs)
    return decorated_function

# Helper function to get projects for an employee
def get_employee_projects_from_csv(employee_id):
    """Get all projects associated with an employee from CSV data"""
    # Convert employee_id to int for comparison
    try:
        employee_id = int(employee_id)
    except (ValueError, TypeError):
        return []
    
    # Find project IDs for this employee
    project_ids = []
    for rel in project_employees_data:
        try:
            rel_emp_id = int(rel.get('EmployeeID', 0))
            if rel_emp_id == employee_id:
                project_id = int(rel.get('ProjectID', 0))
                if project_id:
                    project_ids.append(project_id)
        except (ValueError, TypeError):
            continue
    
    # Find projects matching those IDs
    employee_projects = []
    for proj in projects_data:
        try:
            proj_id = int(proj.get('id', 0))
            if proj_id in project_ids:
                employee_projects.append(proj)
        except (ValueError, TypeError):
            continue
    
    return employee_projects

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/login')
def login():
    """Initiate Azure AD login"""
    if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        # If Azure AD not configured, allow access
        return redirect(url_for('index'))
    try:
        flow = _build_auth_code_flow(scopes=AZURE_SCOPE)
        if not flow:
            return redirect(url_for('index'))
        session["flow"] = flow
        return redirect(session["flow"]["auth_uri"])
    except Exception as e:
        print(f"Error in login: {e}")
        return redirect(url_for('index'))

def _build_auth_code_flow(authority=None, scopes=None):
    """Build authorization code flow"""
    if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        return None
    try:
        cca = _build_msal_app(authority=authority)
        if not cca:
            return None
        return cca.initiate_auth_code_flow(
            scopes or AZURE_SCOPE,
            redirect_uri=url_for("authorized", _external=True)
        )
    except Exception as e:
        print(f"Error building auth code flow: {e}")
        return None

@app.route('/getAToken')
def authorized():
    """Handle Azure AD callback"""
    try:
        cache = _load_cache()
        cca = _build_msal_app(cache=cache)
        if not cca:
            return redirect(url_for('index'))
        
        result = cca.acquire_token_by_auth_code_flow(
            session.get("flow", {}),
            request.args
        )
        _save_cache(cache)
        
        if "error" in result:
            return f"Login failed: {result.get('error_description')}", 400
        
        # Store user info in session
        session["user"] = result.get("id_token_claims")
        return redirect(url_for('index'))
    except Exception as e:
        return f"Login error: {str(e)}", 400

@app.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    # Redirect to Azure AD logout
    if AZURE_AUTHORITY:
        logout_url = f"{AZURE_AUTHORITY}/oauth2/v2.0/logout?post_logout_redirect_uri={url_for('index', _external=True)}"
        return redirect(logout_url)
    return redirect(url_for('index'))

@app.route('/api/practice-areas')
@login_required
def get_practice_areas():
    """Get all unique practice areas (project types) from projects"""
    ensure_data_loaded()
    try:
        practice_areas = set()
        for project in projects_data:
            pa = project.get('practice_area', '').strip()
            if pa:
                practice_areas.add(pa)
        return jsonify({
            "practice_areas": sorted(list(practice_areas))
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sub-practice-areas')
@login_required
def get_sub_practice_areas():
    """Get all unique sub-practice areas from projects"""
    ensure_data_loaded()
    try:
        sub_practice_areas = set()
        for project in projects_data:
            spa = project.get('sub_practice_area', '').strip()
            if spa:
                sub_practice_areas.add(spa)
        return jsonify({
            "sub_practice_areas": sorted(list(sub_practice_areas))
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/user')
def get_user():
    """Get current user info"""
    # If Azure AD is not configured, return not authenticated (but don't error)
    if not AZURE_CLIENT_ID or not AZURE_TENANT_ID:
        return jsonify({"authenticated": False}), 200
    try:
        # Check if user is authenticated
        token = _get_token_from_cache()
        if not token or "error" in token:
            return jsonify({"authenticated": False}), 200
        
        user = session.get("user", {})
        return jsonify({
            "name": user.get("name", "Unknown"),
            "email": user.get("preferred_username", user.get("email", "Unknown")),
            "authenticated": True
        })
    except Exception as e:
        print(f"Error in get_user: {e}")
        return jsonify({"authenticated": False}), 200

@app.route('/api/test')
def test_api():
    """Simple test endpoint to verify API is working"""
    ensure_data_loaded()
    return jsonify({
        "message": "API is working!",
        "employees_loaded": len(employees_data),
        "projects_loaded": len(projects_data),
        "sample_employee": employees_data[0] if employees_data else None
    })

def matches_filter(value, filter_list):
    """Check if a value matches any item in the filter list (partial match)"""
    if not filter_list:
        return True
    return any(str(f).lower() in str(value).lower() for f in filter_list)

def matches_exact_filter(value, filter_list):
    """Check if a value exactly matches any item in the filter list (exact match)"""
    if not filter_list:
        return True
    return any(str(f).lower() == str(value).lower() for f in filter_list)

def matches_range_filter(value, range_filters):
    """Check if a numeric value falls within any of the specified ranges"""
    if not range_filters or value is None:
        return True
    
    for range_filter in range_filters:
        if '-' in range_filter:
            # Handle ranges like "3-5", "6-10"
            try:
                min_val, max_val = map(int, range_filter.split('-'))
                if min_val <= value <= max_val:
                    return True
            except ValueError:
                continue
        elif range_filter.endswith('+'):
            # Handle ranges like "26+"
            try:
                min_val = int(range_filter[:-1])
                if value >= min_val:
                    return True
            except ValueError:
                continue
    
    return False

@app.route('/api/employees')
@login_required
def get_employees():
    print("=" * 60)
    print("API ENDPOINT CALLED: /api/employees")
    print("=" * 60)
    ensure_data_loaded()  # Ensure data is loaded (handles Flask debug reloader)
    try:
        # Get query parameters (support multiple values)
        # Filter out empty strings from filter lists
        def parse_filter_list(param_name):
            value = request.args.get(param_name, '')
            if not value:
                return []
            return [v.strip() for v in value.split(',') if v.strip()]
        
        practice_area = parse_filter_list('practice_area')
        sub_practice_area = parse_filter_list('sub_practice_area')
        region = parse_filter_list('region')
        studio = parse_filter_list('studio')
        years_experience = parse_filter_list('years_experience')
        years_at_pe = parse_filter_list('years_at_pe')
        role = parse_filter_list('role')
        job_title_filter = parse_filter_list('job_title')
        status = parse_filter_list('status')
        name_search = request.args.get('name_search', '').strip()
        
        print("Fetching all employees from CSV data...")
        print(f"DEBUG: employees_data length: {len(employees_data) if employees_data else 0}")
        
        if not employees_data:
            print("DEBUG: employees_data is empty!")
            return jsonify({"employees": [], "message": "No employees found in CSV data."})

        # Process employees from CSV
        results = []
        print(f"Processing {len(employees_data)} employees from CSV...")
        print(f"DEBUG: Filters - practice_area={practice_area}, sub_practice_area={sub_practice_area}, region={region}, studio={studio}, years_experience={years_experience}, years_at_pe={years_at_pe}, role={role}, job_title={job_title_filter}, status={status}, name_search={name_search}")
        
        # Process employees with new filtering capabilities
        employees_processed = 0
        employees_skipped = 0
        for i, emp in enumerate(employees_data):
            if i % 100 == 0:
                print(f"Processed {i}/{len(employees_data)} employees...")
            
            first_name = emp.get("first_name", "N/A")
            last_name = emp.get("last_name", "N/A")
            
            employees_processed += 1
            job_title = emp.get("job_title", "N/A")
            email = emp.get("email", "N/A")
            phone = emp.get("work_phone", "N/A")
            title = emp.get("title", "N/A")
            office = emp.get("studio_office") or emp.get("office", "N/A")
            
            # Handle total_years_in_industry - try to convert to int
            total_years = emp.get("total_years_in_industry")
            try:
                if total_years:
                    total_years = int(float(str(total_years))) if str(total_years).strip() else None
            except (ValueError, TypeError):
                total_years = None
            
            # Handle current_years_with_this_firm - try to convert to int
            employee_years_at_pe = emp.get("current_years_with_this_firm")
            try:
                if employee_years_at_pe:
                    employee_years_at_pe = int(float(str(employee_years_at_pe))) if str(employee_years_at_pe).strip() else None
            except (ValueError, TypeError):
                employee_years_at_pe = None
            
            employee_status = emp.get("status", "Active")  # Employee status

            # Apply filters
            # Years of experience filter - convert to int if possible
            if years_experience:
                try:
                    total_years_int = int(total_years) if total_years else 0
                    if not matches_range_filter(total_years_int, years_experience):
                        continue
                except (ValueError, TypeError):
                    continue  # Skip if can't convert to int
            
            # Years at PE filter - convert to int if possible
            if years_at_pe:
                try:
                    years_at_pe_int = int(employee_years_at_pe) if employee_years_at_pe else 0
                    if not matches_range_filter(years_at_pe_int, years_at_pe):
                        continue
                except (ValueError, TypeError):
                    continue  # Skip if can't convert to int
            
            # Role filter (use exact matching on title only for organizational hierarchy)
            if role and not matches_exact_filter(title, role):
                continue
            
            # Job Title filter (use exact matching on job_title for specific job functions)
            if job_title_filter and not matches_exact_filter(job_title, job_title_filter):
                continue
            
            # Status filter (use exact matching to avoid "Active" matching "Inactive")
            if status and not matches_exact_filter(employee_status, status):
                employees_skipped += 1
                continue

            # Employee passed all initial filters, add to results
            if i < 3:  # Log first 3 employees for debugging
                print(f"DEBUG: Adding employee {i}: {first_name} {last_name} (id={emp.get('id')})")
            
            results.append({
                "id": emp.get("id", 0),
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "title": title,
                "job_title": job_title,
                "office": office if office != "N/A" else "",
                "phone": phone if phone and phone != "N/A" else "",
                "total_years": total_years,
                "years_at_pe": employee_years_at_pe,
                "status": employee_status,
                "img_url": "",  # No images from CSV
                "img_src": "CSV data"
            })

        print(f"DEBUG: Summary - Processed: {employees_processed}, Skipped: {employees_skipped}, Results: {len(results)}")
        print(f"DEBUG: After initial processing, {len(results)} employees in results")
        if len(results) == 0 and len(employees_data) > 0:
            print(f"DEBUG: WARNING - No employees in results but {len(employees_data)} employees loaded!")
            print(f"DEBUG: First employee sample: {employees_data[0] if employees_data else 'N/A'}")
            print(f"DEBUG: First employee fields: {list(employees_data[0].keys())[:10] if employees_data else 'N/A'}")
        
        # Apply name search filter
        if name_search:
            print(f"DEBUG: Applying name search filter: '{name_search}'")
            search_terms = name_search.lower().split()
            filtered_results = []
            for emp in results:
                full_name = f"{emp['first_name']} {emp['last_name']}".lower()
                if any(term in emp['first_name'].lower() or 
                       term in emp['last_name'].lower() or
                       term in full_name
                       for term in search_terms):
                    filtered_results.append(emp)
            results = filtered_results
            print(f"DEBUG: After name search, {len(results)} employees remain")
        
        # Apply studio filter
        if studio:
            print(f"DEBUG: Applying studio filter: {studio}")
            filtered_results = []
            for emp in results:
                emp_office = emp.get('office', '')
                if any(studio_filter.lower() in emp_office.lower() for studio_filter in studio):
                    filtered_results.append(emp)
            results = filtered_results
            print(f"DEBUG: After studio filter, {len(results)} employees remain")
        
        # Apply practice_area, sub_practice_area, and region filters (requires project lookup)
        if practice_area or sub_practice_area or region:
            filtered_results = []
            for emp in results:
                emp_id = emp.get('id')
                # Get projects for this employee
                emp_projects = get_employee_projects_from_csv(emp_id)
                
                # Check practice_area filter
                if practice_area:
                    emp_practice_areas = [p.get('practice_area', '') for p in emp_projects]
                    if not any(pa and any(pa_filter.lower() in pa.lower() for pa_filter in practice_area) 
                              for pa in emp_practice_areas):
                        continue
                
                # Check sub_practice_area filter
                if sub_practice_area:
                    emp_sub_practice_areas = [p.get('sub_practice_area', '') for p in emp_projects]
                    if not any(spa and any(spa_filter.lower() in spa.lower() for spa_filter in sub_practice_area) 
                              for spa in emp_sub_practice_areas):
                        continue
                
                # Check region filter
                if region:
                    emp_regions = [p.get('region', '') for p in emp_projects]
                    if not any(rg and any(rg_filter.lower() in rg.lower() for rg_filter in region) 
                              for rg in emp_regions):
                        continue
                
                filtered_results.append(emp)
            results = filtered_results

        # Sort results
        results.sort(key=lambda r: ((r["last_name"] or "").lower(), (r["first_name"] or "").lower()))
        
        print(f"DEBUG: Final results count: {len(results)}")
        if len(results) > 0:
            print(f"DEBUG: Sample employee: {results[0].get('first_name')} {results[0].get('last_name')}")
        
        # Count how many employees have images
        employees_with_images = sum(1 for emp in results if emp.get("img_url"))
        
        return jsonify({
            "employees": results,
            "total_found": len(results),
            "employees_with_images": employees_with_images,
            "search_method": "csv_data",
            "debug_info": {
                "total_employees_in_system": len(employees_data),
                "data_source": "CSV files"
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error in get_employees: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/api/projects')
@login_required
def get_projects():
    """Get all projects with optional filtering"""
    ensure_data_loaded()
    try:
        # Get filter parameters
        practice_area = request.args.get('practice_area', '')
        sub_practice_area = request.args.get('sub_practice_area', '')
        region = request.args.get('region', '')
        service_type = request.args.get('service_type', '')
        status = request.args.get('status', '')
        
        print(f"Fetching projects from CSV with filters: practice_area={practice_area}, sub_practice_area={sub_practice_area}, region={region}, service_type={service_type}, status={status}")
        
        if not projects_data:
            return jsonify({"projects": [], "message": "No projects found in CSV data."})
        
        # Apply filters from CSV data
        filtered_projects = []
        for project in projects_data:
            # Get project fields from CSV
            project_id = project.get('id', 0)
            project_name = project.get('name', 'Unknown Project')
            project_practice_area = project.get('practice_area', '')
            project_sub_practice_area = project.get('sub_practice_area', '')
            project_region = project.get('region', '')
            # Note: service_type and status may not be in CSV, so we'll skip those filters if not available
            project_status = project.get('status', '')
            
            # Check practice area filter
            if practice_area:
                if not project_practice_area or practice_area.lower() not in project_practice_area.lower():
                    continue
            
            # Check sub_practice_area filter
            if sub_practice_area:
                if not project_sub_practice_area or sub_practice_area.lower() not in project_sub_practice_area.lower():
                    continue
            
            # Check region filter
            if region:
                if not project_region or region.lower() not in project_region.lower():
                    continue
            
            # Check status filter (if available in CSV)
            if status and project_status:
                if status.lower() not in project_status.lower():
                    continue
            
            # Note: service_type may not be in CSV, so we skip that filter for now
            
            filtered_projects.append({
                "id": project_id,
                "name": project_name,
                "practice_area": project_practice_area,
                "sub_practice_area": project_sub_practice_area,
                "region": project_region,
                "service_type": "",  # Not available in CSV
                "status": project_status,
                "openasset_url": f"https://perkinseastman.openasset.com/page/project/{project_id}/employee-roles"
            })
        
        # Sort by name
        filtered_projects.sort(key=lambda p: p["name"].lower())
        
        print(f"Found {len(filtered_projects)} projects after filtering")
        
        return jsonify({
            "projects": filtered_projects,
            "total_found": len(filtered_projects),
            "total_in_system": len(projects_data),
            "filters_applied": {
                "practice_area": practice_area,
                "sub_practice_area": sub_practice_area,
                "region": region,
                "service_type": service_type,
                "status": status
            },
            "data_source": "CSV files"
        })
        
    except Exception as e:
        import traceback
        print(f"Error in get_projects: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/api/employee/<int:employee_id>/projects')
@login_required
def get_employee_projects(employee_id):
    """Get projects associated with a specific employee from CSV data"""
    ensure_data_loaded()
    try:
        print(f"Fetching projects for employee {employee_id} from CSV...")
        
        # Get projects for this employee from CSV
        employee_projects = get_employee_projects_from_csv(employee_id)
        
        if not employee_projects:
            return jsonify({
                "employee_id": employee_id,
                "projects": [],
                "total_projects": 0,
                "message": f"No projects found for employee {employee_id}"
            })
        
        # Format projects for response
        formatted_projects = []
        for project in employee_projects:
            project_id = project.get('id', 0)
            project_name = project.get('name', 'Unknown Project')
            project_practice_area = project.get('practice_area', '')
            project_region = project.get('region', '')
            project_status = project.get('status', '')
            
            formatted_projects.append({
                "id": project_id,
                "name": project_name,
                "practice_area": project_practice_area,
                "sub_practice_area": project.get('sub_practice_area', ''),
                "region": project_region,
                "service_type": "",  # Not available in CSV
                "status": project_status,
                "openasset_url": f"https://perkinseastman.openasset.com/page/project/{project_id}/employee-roles"
            })
        
        # Sort by name
        formatted_projects.sort(key=lambda p: p["name"].lower())
        
        print(f"Found {len(formatted_projects)} projects for employee {employee_id}")
        
        return jsonify({
            "employee_id": employee_id,
            "projects": formatted_projects,
            "total_projects": len(formatted_projects),
            "message": f"Found {len(formatted_projects)} projects via CSV data",
            "data_source": "CSV files"
        })
        
    except Exception as e:
        import traceback
        print(f"Error in get_employee_projects: {e}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route('/api/export/employees')
@login_required
def export_employees_excel():
    """Export filtered employees to Excel file"""
    ensure_data_loaded()
    try:
        # Get the same filters as the main employee search
        # Filter out empty strings from filter lists
        def parse_filter_list(param_name):
            value = request.args.get(param_name, '')
            if not value:
                return []
            return [v.strip() for v in value.split(',') if v.strip()]
        
        practice_area = parse_filter_list('practice_area')
        sub_practice_area = parse_filter_list('sub_practice_area')
        region = parse_filter_list('region')
        studio = parse_filter_list('studio')
        years_experience = parse_filter_list('years_experience')
        years_at_pe = parse_filter_list('years_at_pe')
        role = parse_filter_list('role')
        job_title_filter = parse_filter_list('job_title')
        status = parse_filter_list('status')
        name_search = request.args.get('name_search', '').strip()
        
        print("Exporting employees with filters from CSV...")
        
        if not employees_data:
            return jsonify({"error": "No employees found in CSV data"}), 404

        # Apply the same filtering logic as the main endpoint
        results = []
        for i, emp in enumerate(employees_data):
            if i % 100 == 0:
                print(f"Processing {i}/{len(employees_data)} employees...")
            
            first_name = emp.get("first_name", "N/A")
            last_name = emp.get("last_name", "N/A")
            job_title = emp.get("job_title", "N/A")
            email = emp.get("email", "N/A")
            phone = emp.get("work_phone", "N/A")
            title = emp.get("title", "N/A")
            office = emp.get("studio_office") or emp.get("office", "N/A")
            
            # Handle total_years_in_industry - try to convert to int
            total_years = emp.get("total_years_in_industry")
            try:
                if total_years:
                    total_years = int(float(str(total_years))) if str(total_years).strip() else None
            except (ValueError, TypeError):
                total_years = None
            
            # Handle current_years_with_this_firm - try to convert to int
            employee_years_at_pe = emp.get("current_years_with_this_firm")
            try:
                if employee_years_at_pe:
                    employee_years_at_pe = int(float(str(employee_years_at_pe))) if str(employee_years_at_pe).strip() else None
            except (ValueError, TypeError):
                employee_years_at_pe = None
            
            employee_status = emp.get("status", "Active")

            # Apply the same filters as the main endpoint
            # Name search filter
            if name_search:
                name_terms = name_search.lower().split()
                full_name = f"{first_name} {last_name}".lower()
                if not all(term in full_name for term in name_terms):
                    continue
            
            # Studio filter
            if studio:
                if not any(studio_filter.lower() in office.lower() for studio_filter in studio if office):
                    continue
            
            # Years of experience filter
            if years_experience:
                try:
                    total_years_int = int(total_years) if total_years else 0
                    if not matches_range_filter(total_years_int, years_experience):
                        continue
                except (ValueError, TypeError):
                    continue
            
            # Years at PE filter
            if years_at_pe:
                try:
                    years_at_pe_int = int(employee_years_at_pe) if employee_years_at_pe else 0
                    if not matches_range_filter(years_at_pe_int, years_at_pe):
                        continue
                except (ValueError, TypeError):
                    continue
            
            # Role filter (title)
            if role and not matches_exact_filter(title, role):
                continue
            
            # Job Title filter
            if job_title_filter and not matches_exact_filter(job_title, job_title_filter):
                continue
            
            # Status filter
            if status and not matches_exact_filter(employee_status, status):
                continue
            
            # Apply practice_area, sub_practice_area, and region filters (requires project lookup)
            if practice_area or sub_practice_area or region:
                emp_id = emp.get('id')
                emp_projects = get_employee_projects_from_csv(emp_id)
                
                if practice_area:
                    emp_practice_areas = [p.get('practice_area', '') for p in emp_projects]
                    if not any(pa and any(pa_filter.lower() in pa.lower() for pa_filter in practice_area) 
                              for pa in emp_practice_areas):
                        continue
                
                if sub_practice_area:
                    emp_sub_practice_areas = [p.get('sub_practice_area', '') for p in emp_projects]
                    if not any(spa and any(spa_filter.lower() in spa.lower() for spa_filter in sub_practice_area) 
                              for spa in emp_sub_practice_areas):
                        continue
                
                if region:
                    emp_regions = [p.get('region', '') for p in emp_projects]
                    if not any(rg and any(rg_filter.lower() in rg.lower() for rg_filter in region) 
                              for rg in emp_regions):
                        continue

            # Add to results
            results.append({
                "Status": employee_status,
                "Studio": office if office != "N/A" else "",
                "Location": office if office != "N/A" else "",
                "Title": title,
                "Role": job_title,
                "Name": f"{first_name} {last_name}",
                "Years of Experience": total_years if total_years else "0",
                "Years at PE": employee_years_at_pe if employee_years_at_pe else "0"
            })

        print(f"Exporting {len(results)} employees to Excel")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        
        # Define headers
        headers = ["Status", "Studio", "Location", "Title", "Role", "Name", "Years of Experience", "Years at PE"]
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, employee in enumerate(results, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=employee.get(header, ""))
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, len(results) + 2):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create filename with timestamp
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"perkins_eastman_employees_{timestamp}.xlsx"
        
        # Save to memory buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exporting employees: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
